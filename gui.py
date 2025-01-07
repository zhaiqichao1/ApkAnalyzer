import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
import os
import json
from apk_analyzer import ApkAnalyzer
from datetime import datetime, timezone, timedelta
from database import Database
from output_handler import OutputHandler
import subprocess
import sqlite3
import openpyxl

class ApkAnalyzerGUI:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("APK网络分析工具")
        
        # 设置窗口最小尺寸和默认状态
        self.root.minsize(1024, 768)
        self.root.state('zoomed')
        
        # 配置根窗口的网格权重
        self.root.grid_rowconfigure(0, weight=1)
        self.root.grid_columnconfigure(0, weight=1)
        
        self.analyzer = ApkAnalyzer()
        
        # 初始化变量
        if self.analyzer.nox_path:
            self.nox_path_var = tk.StringVar(value=self.analyzer.nox_path)
        else:
            self.nox_path_var = tk.StringVar()
        
        self.apk_path_var = tk.StringVar()
        self.path_status_var = tk.StringVar()
        self.mode_var = tk.StringVar(value="single")
        
        if self.analyzer.nox_path:
            self.path_status_var.set("✓ 已加载保存的路径")
        
        self.request_set = set()
        self.selected_apks = []
        
        # 初始化数据库
        try:
            from database import Database
            self.db = Database()
        except Exception as e:
            print(f"初始化数据库时出错: {str(e)}")
            self.db = None
        
        self.setup_gui()
        
    def setup_gui(self):
        # 创建主框架
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky='nsew')
        
        # 创建左右分隔的面板
        panel = ttk.PanedWindow(main_frame, orient=tk.HORIZONTAL)
        panel.grid(row=0, column=0, sticky='nsew')
        main_frame.grid_rowconfigure(0, weight=1)
        main_frame.grid_columnconfigure(0, weight=1)
        
        # 左侧面板
        left_frame = ttk.Frame(panel)
        panel.add(left_frame, weight=1)  # 修改权重为1
        left_frame.grid_columnconfigure(0, weight=1)
        left_frame.grid_rowconfigure(1, weight=1)
        
        # 控制区域
        control_frame = ttk.LabelFrame(left_frame, text="控制面板", padding="5")
        control_frame.grid(row=0, column=0, columnspan=4, sticky='ew', pady=5)
        
        # 夜神模拟器路径设置
        ttk.Label(control_frame, text="夜神模拟器路径:").grid(row=0, column=0, sticky='w', padx=5)
        ttk.Entry(control_frame, textvariable=self.nox_path_var, width=50).grid(row=0, column=1, padx=5)
        ttk.Button(control_frame, text="浏览", command=self.browse_nox_path).grid(row=0, column=2, padx=5)
        ttk.Button(control_frame, text="清除", command=self.clear_saved_path).grid(row=0, column=3, padx=5)
        
        # 路径状态显示
        ttk.Label(control_frame, textvariable=self.path_status_var).grid(row=1, column=1, sticky='w', padx=5)
        
        # APK文件选择
        ttk.Label(control_frame, text="APK路径:").grid(row=2, column=0, sticky='w', pady=10)
        ttk.Entry(control_frame, textvariable=self.apk_path_var, width=50).grid(row=2, column=1, padx=5)
        ttk.Button(control_frame, text="选择单个APK", command=self.browse_apk_file).grid(row=2, column=2)
        ttk.Button(control_frame, text="选择多个APK", command=self.browse_multiple_apks).grid(row=2, column=3)
        
        # 操作按钮
        button_frame = ttk.Frame(control_frame)
        button_frame.grid(row=3, column=0, columnspan=4, pady=10)
        
        ttk.Button(button_frame, text="开始分析", command=self.start_analysis).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="查看数据库", command=self.show_database_view).pack(side=tk.LEFT, padx=5)
        
        # 在控制面板添加 frida-server 操作按钮
        frida_frame = ttk.Frame(control_frame)
        frida_frame.grid(row=4, column=0, columnspan=4, pady=5)
        
        ttk.Button(frida_frame, text="推送 frida-server", 
                   command=self.push_frida_server).pack(side=tk.LEFT, padx=5)
        ttk.Button(frida_frame, text="验证 frida-server", 
                   command=self.verify_frida_server).pack(side=tk.LEFT, padx=5)
        
        # 网络请求数据表格
        request_frame = ttk.LabelFrame(left_frame, text="网络请求数据", padding="5")
        request_frame.grid(row=1, column=0, sticky='nsew', padx=5, pady=5)
        request_frame.grid_rowconfigure(0, weight=1)
        request_frame.grid_columnconfigure(0, weight=1)
        
        # 创建表格和滚动条
        table_frame = ttk.Frame(request_frame)
        table_frame.grid(row=0, column=0, sticky='nsew')
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)
        
        # 创建网络请求表格
        columns = ('时间', '域名', 'IP地址', '端口', '国家', '地区', '城市', 'ISP', '组织')
        self.tree = ttk.Treeview(table_frame, columns=columns, show='headings')
        
        # 设置列标题和宽度
        column_widths = {
            '时间': 150, '域名': 200, 'IP地址': 120, '端口': 80,
            '国家': 100, '地区': 100, '城市': 100, 'ISP': 150, '组织': 150
        }
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=column_widths[col])
        
        # 添加滚动条
        y_scroll = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.tree.yview)
        x_scroll = ttk.Scrollbar(table_frame, orient=tk.HORIZONTAL, command=self.tree.xview)
        self.tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)
        
        # 表格布局
        self.tree.grid(row=0, column=0, sticky='nsew')
        y_scroll.grid(row=0, column=1, sticky='ns')
        x_scroll.grid(row=1, column=0, sticky='ew')
        
        # 创建右键菜单
        self.tree_menu = tk.Menu(self.root, tearoff=0)
        self.tree_menu.add_command(label="复制选中行", command=lambda: self.copy_selected_rows(self.tree))
        self.tree_menu.add_command(label="复制所有数据", command=lambda: self.copy_all_rows(self.tree))
        
        # 绑定右键菜单
        self.tree.bind("<Button-3>", self.show_menu)
        
        # 底部区域
        bottom_frame = ttk.Frame(main_frame)
        bottom_frame.grid(row=1, column=0, sticky='ew', pady=(5,0))
        bottom_frame.grid_columnconfigure(0, weight=1)
        
        # 日志区域
        log_frame = ttk.LabelFrame(bottom_frame, text="运行日志", padding="5")
        log_frame.grid(row=0, column=0, sticky='ew', padx=5)
        log_frame.grid_columnconfigure(0, weight=1)
        
        self.log_text = tk.Text(log_frame, height=6,
                               font=('Consolas', 9),
                               bg='#f8f9fa',
                               wrap=tk.WORD)
        self.log_text.grid(row=0, column=0, sticky='ew', padx=5, pady=5)
        
        # 进度条
        self.progress_var = tk.DoubleVar()
        self.progress = ttk.Progressbar(bottom_frame,
                                      style='Horizontal.TProgressbar',
                                      mode='determinate',
                                      variable=self.progress_var)
        self.progress.grid(row=1, column=0, sticky='ew', padx=5, pady=5)
        
    def browse_nox_path(self):
        path = filedialog.askdirectory(title="选择夜神模拟器安装目录")
        if path:
            if self.analyzer.set_nox_path(path):
                self.nox_path_var.set(path)
                self.path_status_var.set("✓ 路径已保存")
            else:
                self.path_status_var.set("✗ 无效的路径")
            
    def browse_apk_file(self):
        path = filedialog.askopenfilename(title="选择APK文件", 
                                        filetypes=[("APK文件", "*.apk")])
        if path:
            self.apk_path_var.set(path)
            self.mode_var.set("single")

    def browse_multiple_apks(self):
        """选择多个APK文件"""
        paths = filedialog.askopenfilenames(
            title="选择APK文件",
            filetypes=[("APK文件", "*.apk")]
        )
        if paths:
            # 将选择的文件路径保存为列表
            self.selected_apks = list(paths)
            # 显示选择的文件数量
            self.apk_path_var.set(f"已选择 {len(self.selected_apks)} 个APK文件")
            self.mode_var.set("batch")
            
            # 显示选择的文件列表
            file_list = "\n".join([os.path.basename(p) for p in self.selected_apks])
            self.log(f"已选择以下文件：\n{file_list}\n")

    def log(self, message):
        self.log_text.insert(tk.END, f"{message}\n")
        self.log_text.see(tk.END)
        self.root.update()
        
    def start_analysis(self):
        # 检查输入
        nox_path = self.nox_path_var.get().strip()
        
        if not nox_path:
            messagebox.showerror("错误", "请设置夜神模拟器路径")
            return
        
        if not hasattr(self, 'selected_apks') and not self.apk_path_var.get().strip():
            messagebox.showerror("错误", "请选择APK文件")
            return
        
        # 清空日志
        self.log_text.delete(1.0, tk.END)
        
        # 在新线程中运行分析
        thread = threading.Thread(target=self.run_analysis)
        thread.daemon = True
        thread.start()
        
    def run_analysis(self):
        try:
            # 设置夜神模拟器路径
            if not self.analyzer.set_nox_path(self.nox_path_var.get()):
                self.log("错误: 无效的夜神模拟器路径")
                return
                
            # 设置环境
            self.log("正在设置环境...")
            if not self.analyzer.setup_environment():
                self.log("环境设置失败")
                return
                
            # 连接模拟器
            self.log("正在连接模拟器...")
            if not self.analyzer.connect_emulator():
                self.log("模拟器连接失败")
                return
                
            # 清空表格
            self.clear_table()
            
            # 设置回调
            self.analyzer.set_progress_callback(self.update_table)
            
            # 根据是否有选中的多个文件来决定分析模式
            if hasattr(self, 'selected_apks') and self.selected_apks:
                self.analyze_batch()
            else:
                self.analyze_single()
                
        except Exception as e:
            self.log(f"错误: {str(e)}")
            messagebox.showerror("错误", str(e))
            
    def analyze_single(self):
        apk_path = self.apk_path_var.get()
        if not os.path.exists(apk_path):
            self.log("错误: APK文件不存在")
            messagebox.showerror("错误", "APK文件不存在")
            return
            
        self.log(f"开始分析: {os.path.basename(apk_path)}")
        self.progress_var.set(0)
        
        try:
            results = self.analyzer.analyze_apk(apk_path)
            if results:
                # 只保存到数据库
                self.save_analysis_results(apk_path, results, "成功")
                
                self.log("分析完成，结果已保存")
                messagebox.showinfo("成功", 
                    f"APK {os.path.basename(apk_path)} 分析完成\n"
                    f"发现 {len(results['requests'])} 个网络请求"
                )
            else:
                error_msg = "分析失败：未能获取结果"
                self.log(error_msg)
                self.save_analysis_results(apk_path, None, f"失败: {error_msg}")
                messagebox.showerror("错误", error_msg)
                
        except Exception as e:
            error_msg = str(e)
            self.log(f"分析出错: {error_msg}")
            self.save_analysis_results(apk_path, None, f"失败: {error_msg}")
            messagebox.showerror("错误", error_msg)
            
        self.progress_var.set(100)
        
    def analyze_batch(self):
        """批量分析APK"""
        if not hasattr(self, 'selected_apks') or not self.selected_apks:
            self.log("错误: 未选择APK文件")
            messagebox.showerror("错误", "未选择APK文件")
            return
        
        total = len(self.selected_apks)
        success_count = 0
        failed_count = 0
        failed_files = []
        
        # 更新进度条最大值
        self.progress_var.set(0)
        
        for i, apk_path in enumerate(self.selected_apks, 1):
            try:
                # 更新进度显示
                progress = (i-1) * 100 / total
                self.progress_var.set(progress)
                self.log(f"\n正在分析 ({i}/{total}): {os.path.basename(apk_path)}")
                
                if not os.path.exists(apk_path):
                    error_msg = f"文件不存在: {apk_path}"
                    self.log(error_msg)
                    failed_count += 1
                    failed_files.append((os.path.basename(apk_path), error_msg))
                    continue
                
                # 分析APK
                results = self.analyzer.analyze_apk(apk_path)
                if results:
                    # 保存结果
                    self.save_analysis_results(apk_path, results, "成功")
                    success_count += 1
                    self.log(f"分析完成: {os.path.basename(apk_path)}")
                else:
                    error_msg = "分析失败：未能获取结果"
                    self.log(error_msg)
                    failed_count += 1
                    failed_files.append((os.path.basename(apk_path), error_msg))
                
                # 每分析10个文件后重启ADB服务
                if i % 10 == 0:
                    self.log("\n重启ADB服务...")
                    try:
                        self.analyzer.connect_emulator()
                    except:
                        pass
                
            except Exception as e:
                error_msg = str(e)
                self.log(f"分析出错: {error_msg}")
                failed_count += 1
                failed_files.append((os.path.basename(apk_path), error_msg))
                continue
        
        # 更新最终进度
        self.progress_var.set(100)
        
        # 显示分析结果
        summary = (
            f"批量分析完成：\n\n"
            f"总计：{total} 个文件\n"
            f"成功：{success_count} 个\n"
            f"失败：{failed_count} 个\n\n"
        )
        
        if failed_files:
            summary += "失败的文件：\n"
            for filename, error in failed_files:
                summary += f"{filename}: {error}\n"
        
        self.log("\n" + summary)
        
        if failed_count > 0:
            messagebox.showwarning("分析结果", summary)
        else:
            messagebox.showinfo("分析结果", summary)
        
    def clear_saved_path(self):
        self.nox_path_var.set("")
        self.analyzer.nox_path = ""
        self.analyzer.adb_path = None
        self.analyzer.config['nox_path'] = ""
        self.analyzer.save_config()
        self.path_status_var.set("已清除保存的路径")
        
    def update_table(self, request_info):
        """更新表格数据"""
        try:
            # 生成唯一标识（使用域名和IP的组合）
            unique_id = f"{request_info.get('domain', '')}-{request_info.get('ip', '')}"
            
            # 如果已经显示过该请求，则跳过
            if unique_id in self.request_set:
                return
            
            # 添加到已显示集合
            self.request_set.add(unique_id)
            
            # 处理时间戳
            try:
                timestamp_str = request_info.get('timestamp', '')
                china_tz = timezone(timedelta(hours=8))  # 中国时区 UTC+8
                
                if isinstance(timestamp_str, str):
                    if 'T' in timestamp_str:  # ISO格式
                        if '+' in timestamp_str:
                            dt = datetime.fromisoformat(timestamp_str)
                        else:
                            dt = datetime.fromisoformat(timestamp_str.replace('Z', '+00:00'))
                        dt = dt.astimezone(china_tz)
                        timestamp_str = dt.strftime('%Y-%m-%d %H:%M:%S')
                    elif timestamp_str.isdigit():  # Unix时间戳
                        dt = datetime.fromtimestamp(int(timestamp_str)/1000, timezone.utc)
                        dt = dt.astimezone(china_tz)
                        timestamp_str = dt.strftime('%Y-%m-%d %H:%M:%S')
                    elif '.' in timestamp_str:  # 带毫秒的时间戳
                        try:
                            dt = datetime.strptime(timestamp_str, '%Y-%m-%d %H:%M:%S.%f')
                            dt = dt.replace(tzinfo=timezone.utc)
                            dt = dt.astimezone(china_tz)
                            timestamp_str = dt.strftime('%Y-%m-%d %H:%M:%S')
                        except:
                            # 如果解析失败，尝试其他格式
                            pass
                elif isinstance(timestamp_str, (int, float)):  # 数字类型时间戳
                    dt = datetime.fromtimestamp(timestamp_str/1000, timezone.utc)
                    dt = dt.astimezone(china_tz)
                    timestamp_str = dt.strftime('%Y-%m-%d %H:%M:%S')
                else:
                    timestamp_str = datetime.now(china_tz).strftime('%Y-%m-%d %H:%M:%S')
            except Exception as e:
                print(f"处理时间戳出错: {str(e)}, 原始时间戳: {timestamp_str}")
                timestamp_str = datetime.now(china_tz).strftime('%Y-%m-%d %H:%M:%S')
            
            # 插入数据到表格
            row = self.tree.insert('', 'end', values=(
                timestamp_str,
                request_info.get('domain', ''),
                request_info.get('ip', 'Unknown'),
                request_info.get('port', ''),
                request_info.get('country', 'Unknown'),
                request_info.get('region', 'Unknown'),
                request_info.get('city', 'Unknown'),
                request_info.get('isp', 'Unknown'),
                request_info.get('org', 'Unknown')
            ))
            
            # 自动滚动到最新行
            self.tree.yview_moveto(1)
            self.root.update()
            
        except Exception as e:
            print(f"更新界面时出错: {str(e)}")
            print(f"请求信息: {request_info}")
        
    def clear_table(self):
        """清空表格数据"""
        self.request_set.clear()  # 清空请求集合
        for item in self.tree.get_children():
            self.tree.delete(item)
        
    def save_analysis_results(self, apk_path, results, status):
        """保存分析结果"""
        try:
            # 保存到数据库
            if results:
                try:
                    self.db.save_results(results)
                except Exception as e:
                    print(f"保存到数据库时出错: {str(e)}")
                    raise
            
            # 如果有网络请求数据，自动保存Excel
            if results.get('requests'):
                try:
                    # 获取APK所在目录
                    apk_dir = os.path.dirname(apk_path)
                    apk_name = os.path.splitext(os.path.basename(apk_path))[0]
                    
                    # 构建Excel文件名
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                    excel_name = f"{apk_name}_网络请求_{timestamp}.xlsx"
                    excel_path = os.path.join(apk_dir, excel_name)
                    
                    # 使用OutputHandler保存Excel
                    from output_handler import OutputHandler
                    output_handler = OutputHandler(apk_dir)
                    output_handler.save_results(results)
                    
                    self.log(f"分析结果已保存到: {excel_path}")
                    
                except Exception as e:
                    self.log(f"保存Excel文件时出错: {str(e)}")
                    raise
    
        except Exception as e:
            self.log(f"保存分析结果时出错: {str(e)}")
            raise

    def run(self):
        try:
            self.root.mainloop()
        except Exception as e:
            # 直接显示错误消息
            messagebox.showerror("错误", f"程序运行出错: {str(e)}")

    def show_menu(self, event):
        """显示右键菜单"""
        try:
            self.tree_menu.tk_popup(event.x_root, event.y_root)
        finally:
            self.tree_menu.grab_release()

    def copy_selected_rows(self, tree):
        """复制选中的行"""
        selected_items = tree.selection()
        if not selected_items:
            return
        
        # 获取列标题
        headers = [tree.heading(col)['text'] for col in tree['columns']]
        
        # 准备复制的数据
        rows = [headers]  # 首先添加表头
        for item in selected_items:
            row = list(tree.item(item)['values'])
            rows.append(row)
        
        # 转换为制表符分隔的文本
        text = '\n'.join(['\t'.join(map(str, row)) for row in rows])
        
        # 复制到剪贴板
        self.root.clipboard_clear()
        self.root.clipboard_append(text)
        messagebox.showinfo("提示", f"已复制 {len(selected_items)} 行数据到剪贴板")

    def copy_all_rows(self, tree):
        """复制所有行"""
        # 获取列标题
        headers = [tree.heading(col)['text'] for col in tree['columns']]
        
        # 准备复制的数据
        rows = [headers]  # 首先添加表头
        for item in tree.get_children():
            row = list(tree.item(item)['values'])
            rows.append(row)
        
        # 转换为制表符分隔的文本
        text = '\n'.join(['\t'.join(map(str, row)) for row in rows])
        
        # 复制到剪贴板
        self.root.clipboard_clear()
        self.root.clipboard_append(text)
        messagebox.showinfo("提示", f"已复制 {len(rows)-1} 行数据到剪贴板")

    def export_to_excel(self):
        """导出数据到Excel"""
        try:
            # 确保有选中的记录
            if not self.db_tree.selection():
                messagebox.showwarning("警告", "请先选择要导出的记录")
                return
            
            # 获取选中的记录
            selected_items = self.db_tree.selection()
            
            # 获取保存路径
            file_path = filedialog.asksaveasfilename(
                defaultextension='.xlsx',
                filetypes=[('Excel 文件', '*.xlsx')],
                initialfile='网络请求分析结果.xlsx'
            )
            
            if not file_path:
                return
            
            # 创建Excel工作簿
            wb = openpyxl.Workbook()
            
            # 创建概览sheet
            ws_overview = wb.active
            ws_overview.title = "分析概览"
            
            # 设置概览表头
            overview_headers = ['APK文件名', '分析时间', '请求总数', '域名数量', 'IP数量']
            for col, header in enumerate(overview_headers, 1):
                cell = ws_overview.cell(row=1, column=col)
                cell.value = header
                cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = openpyxl.styles.PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            
            # 获取所有选中记录的详细信息
            row = 2
            for item in selected_items:
                record = self.db_tree.item(item)['values']
                analysis_id = None
                
                # 从数据库获取分析ID
                records = self.db.get_analysis_list()
                for rec in records:
                    if rec[1] == record[1] and str(rec[3]) == record[0]:
                        analysis_id = rec[0]
                        break
                
                if analysis_id:
                    # 获取请求数据
                    requests = self.db.get_requests_by_analysis(analysis_id)
                    
                    # 统计信息
                    domains = len(set(req[0] for req in requests if req[0]))  # 域名列
                    ips = len(set(req[1] for req in requests if req[1]))     # IP列
                    
                    # 写入概览数据
                    ws_overview.cell(row=row, column=1, value=record[1])  # APK文件名
                    ws_overview.cell(row=row, column=2, value=record[0])  # 分析时间
                    ws_overview.cell(row=row, column=3, value=len(requests))  # 请求总数
                    ws_overview.cell(row=row, column=4, value=domains)  # 域名数量
                    ws_overview.cell(row=row, column=5, value=ips)  # IP数量
                    
                    # 创建详细信息sheet
                    sheet_name = f"详细信息_{row-1}"
                    ws_details = wb.create_sheet(sheet_name)
                    
                    # 设置详细信息表头
                    detail_headers = [
                        '时间', '请求类型', '子类型', '域名', 'IP地址', '端口',
                        'URL', '请求方法', '协议', '国家', '地区', '城市',
                        'ISP', '组织'
                    ]
                    
                    # 设置列宽
                    column_widths = {
                        '时间': 20, '请求类型': 15, '子类型': 15, '域名': 30,
                        'IP地址': 15, '端口': 10, 'URL': 50, '请求方法': 10,
                        '协议': 10, '国家': 15, '地区': 15, '城市': 15,
                        'ISP': 20, '组织': 20
                    }
                    
                    # 写入表头并设置样式
                    for col, header in enumerate(detail_headers, 1):
                        cell = ws_details.cell(row=1, column=col)
                        cell.value = header
                        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
                        cell.fill = openpyxl.styles.PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                        cell.alignment = openpyxl.styles.Alignment(horizontal="center")
                        ws_details.column_dimensions[openpyxl.utils.get_column_letter(col)].width = column_widths[header]
                    
                    # 写入详细数据
                    for req_row, request in enumerate(requests, 2):
                        for col, value in enumerate(request, 1):
                            cell = ws_details.cell(row=req_row, column=col)
                            cell.value = str(value) if value is not None else ''
                            cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)
                    
                    row += 1
            
            # 保存Excel文件
            wb.save(file_path)
            messagebox.showinfo("成功", f"数据已导出到:\n{file_path}")
            
        except Exception as e:
            messagebox.showerror("错误", f"导出数据失败: {str(e)}")

    def delete_selected_history(self):
        """删除选中的历史记录"""
        selected = self.history_tree.selection()
        if not selected:
            return
        
        if messagebox.askyesno("确认", "确定要删除选中的记录吗？"):
            for item in selected:
                values = self.history_tree.item(item)['values']
                # 从历史记录列表中删除
                self.history = [h for h in self.history if not (
                    h['time'] == values[0] and 
                    h['filename'] == values[1]
                )]
                # 从显示中删除
                self.history_tree.delete(item)
            
            # 保存更新后的历史记录
            try:
                with open('analysis_history.json', 'w', encoding='utf-8') as f:
                    json.dump(self.history, f, indent=2, ensure_ascii=False)
            except Exception as e:
                self.log(f"保存历史记录失败: {str(e)}")

    def clear_all_history(self):
        """清空所有历史记录"""
        if messagebox.askyesno("确认", "确定要清空所有历史记录吗？"):
            # 清空历史记录列表
            self.history = []
            # 清空显示
            for item in self.history_tree.get_children():
                self.history_tree.delete(item)
            # 保存空的历史记录
            try:
                with open('analysis_history.json', 'w', encoding='utf-8') as f:
                    json.dump(self.history, f, indent=2, ensure_ascii=False)
            except Exception as e:
                self.log(f"保存历史记录失败: {str(e)}")

    def select_apk(self):
        """选择并分析单个APK"""
        try:
            # 获取APK文件路径
            file_path, _ = QFileDialog.getOpenFileName(self, '选择APK文件', 
                                                 self.analyzer.config.get('last_apk_path', ''),
                                                 'APK文件 (*.apk)')
            
            if not file_path:
                return
            
            # 保存最后使用的路径
            self.analyzer.config['last_apk_path'] = file_path
            self.analyzer.save_config()
            
            # 清空表格
            self.table.setRowCount(0)
            
            # 开始分析
            results = self.analyzer.analyze_apk(file_path)
            
            if results:
                # 保存结果
                self.save_analysis_results(results)
                QMessageBox.information(self, '完成', '分析完成，结果已保存')
            
        except Exception as e:
            QMessageBox.critical(self, '错误', f'分析APK时出错: {str(e)}')

    def show_database_view(self):
        """显示数据库查看窗口"""
        try:
            # 创建新窗口
            self.db_window = tk.Toplevel(self.root)
            self.db_window.title("数据库查看")
            self.db_window.geometry("1000x700")
            
            # 创建搜索框架
            search_frame = ttk.LabelFrame(self.db_window, text="搜索条件", padding="5")
            search_frame.pack(fill=tk.X, padx=5, pady=5)
            
            # 文件名搜索
            ttk.Label(search_frame, text="文件名:").grid(row=0, column=0, padx=5)
            self.filename_var = tk.StringVar()
            ttk.Entry(search_frame, textvariable=self.filename_var, width=30).grid(row=0, column=1, padx=5)
            
            # 时间范围
            ttk.Label(search_frame, text="时间范围:").grid(row=0, column=2, padx=5)
            self.date_from_var = tk.StringVar()
            self.date_to_var = tk.StringVar()
            ttk.Entry(search_frame, textvariable=self.date_from_var, width=20).grid(row=0, column=3, padx=5)
            ttk.Label(search_frame, text="至").grid(row=0, column=4)
            ttk.Entry(search_frame, textvariable=self.date_to_var, width=20).grid(row=0, column=5, padx=5)
            
            # 搜索按钮
            ttk.Button(search_frame, text="搜索", command=self.search_database).grid(row=0, column=6, padx=5)
            ttk.Button(search_frame, text="重置", command=self.reset_search).grid(row=0, column=7, padx=5)
            
            # 创建表格框架
            table_frame = ttk.Frame(self.db_window)
            table_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
            
            # 创建表格
            columns = ('时间', '文件名', '请求数', 'Hash值')
            self.db_tree = ttk.Treeview(table_frame, columns=columns, show='headings', selectmode='extended')
            
            # 设置列标题和宽度
            widths = {'时间': 150, '文件名': 300, '请求数': 100, 'Hash值': 250}
            for col in columns:
                self.db_tree.heading(col, text=col, command=lambda c=col: self.sort_treeview(c))
                self.db_tree.column(col, width=widths[col])
            
            # 添加滚动条
            y_scroll = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.db_tree.yview)
            x_scroll = ttk.Scrollbar(table_frame, orient=tk.HORIZONTAL, command=self.db_tree.xview)
            self.db_tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)
            
            # 布局
            self.db_tree.grid(row=0, column=0, sticky='nsew')
            y_scroll.grid(row=0, column=1, sticky='ns')
            x_scroll.grid(row=1, column=0, sticky='ew')
            
            table_frame.grid_rowconfigure(0, weight=1)
            table_frame.grid_columnconfigure(0, weight=1)
            
            # 分页控件
            page_frame = ttk.Frame(self.db_window)
            page_frame.pack(fill=tk.X, padx=5, pady=5)
            
            self.page_size = 100  # 每页显示数量
            self.current_page = 1
            self.total_pages = 1
            
            ttk.Button(page_frame, text="上一页", command=self.prev_page).pack(side=tk.LEFT, padx=5)
            self.page_label = ttk.Label(page_frame, text="1/1")
            self.page_label.pack(side=tk.LEFT, padx=5)
            ttk.Button(page_frame, text="下一页", command=self.next_page).pack(side=tk.LEFT, padx=5)
            
            # 加载初始数据
            self.load_database_page()
            
            # 添加右键菜单
            self.db_menu = tk.Menu(self.db_window, tearoff=0)
            self.db_menu.add_command(label="查看详情", command=self.show_record_details)
            self.db_menu.add_command(label="导出Excel", command=self.export_selected_records)
            self.db_menu.add_separator()
            self.db_menu.add_command(label="删除选中", command=self.delete_selected_records)
            
            self.db_tree.bind('<Button-3>', self.show_db_menu)
            
        except Exception as e:
            messagebox.showerror("错误", f"显示数据库视图失败: {str(e)}")

    def load_database_page(self):
        """加载当前页的数据"""
        try:
            # 清空表格
            for item in self.db_tree.get_children():
                self.db_tree.delete(item)
            
            # 构建查询条件
            conditions = {}
            if self.filename_var.get():
                conditions['filename'] = self.filename_var.get()
            if self.date_from_var.get():
                conditions['date_from'] = self.date_from_var.get()
            if self.date_to_var.get():
                conditions['date_to'] = self.date_to_var.get()
            
            # 获取总记录数
            total_records = self.db.get_total_records(conditions)
            self.total_pages = (total_records + self.page_size - 1) // self.page_size
            
            # 更新页码显示
            self.page_label.config(text=f"{self.current_page}/{self.total_pages}")
            
            # 获取当前页数据
            offset = (self.current_page - 1) * self.page_size
            records = self.db.get_records_page(offset, self.page_size, conditions)
            
            # 填充数据
            for record in records:
                self.db_tree.insert('', 'end', values=record)
            
        except Exception as e:
            messagebox.showerror("错误", f"加载数据失败: {str(e)}")

    def search_database(self):
        """搜索数据库"""
        self.current_page = 1
        self.load_database_page()

    def reset_search(self):
        """重置搜索条件"""
        self.filename_var.set("")
        self.date_from_var.set("")
        self.date_to_var.set("")
        self.search_database()

    def prev_page(self):
        """上一页"""
        if self.current_page > 1:
            self.current_page -= 1
            self.load_database_page()

    def next_page(self):
        """下一页"""
        if self.current_page < self.total_pages:
            self.current_page += 1
            self.load_database_page()

    def sort_treeview(self, col):
        """排序表格"""
        try:
            # 获取所有数据
            data = [(self.db_tree.set(item, col), item) for item in self.db_tree.get_children('')]
            
            # 获取当前排序方向
            if not hasattr(self, '_sort_reverse'):
                self._sort_reverse = {}
            
            # 切换排序方向
            self._sort_reverse[col] = not self._sort_reverse.get(col, False)
            
            # 排序
            data.sort(reverse=self._sort_reverse[col])
            
            # 重新排列项目
            for index, (_, item) in enumerate(data):
                self.db_tree.move(item, '', index)
                
            # 更新表头显示排序方向
            for header in self.db_tree['columns']:
                if header == col:
                    # 显示排序方向指示器
                    direction = '↓' if self._sort_reverse[col] else '↑'
                    self.db_tree.heading(header, text=f"{header} {direction}")
                else:
                    # 移除其他列的排序指示器
                    self.db_tree.heading(header, text=header.rstrip(' ↑↓'))
                
        except Exception as e:
            print(f"排序时出错: {str(e)}")

    def show_record_details(self):
        """显示记录详情"""
        try:
            # 确保有选中的项目
            if not self.db_tree.selection():
                messagebox.showwarning("警告", "请先选择一条记录")
                return
            
            # 获取选中记录
            selected_item = self.db_tree.selection()[0]
            record = self.db_tree.item(selected_item)['values']
            
            # 创建新窗口
            detail_window = tk.Toplevel(self.db_window)
            detail_window.title(f"记录详情 - {record[1]}")
            detail_window.geometry("1000x600")
            
            # 创建表格
            columns = ('域名', 'IP地址', '端口', '时间', '国家', '地区', '城市', 'ISP', '组织')
            tree = ttk.Treeview(detail_window, columns=columns, show='headings')
            
            # 设置列标题和宽度
            widths = {
                '域名': 200, 'IP地址': 120, '端口': 80, '时间': 150,
                '国家': 100, '地区': 100, '城市': 100, 'ISP': 150, '组织': 150
            }
            for col in columns:
                tree.heading(col, text=col)
                tree.column(col, width=widths[col])
            
            # 添加滚动条
            y_scroll = ttk.Scrollbar(detail_window, orient=tk.VERTICAL, command=tree.yview)
            x_scroll = ttk.Scrollbar(detail_window, orient=tk.HORIZONTAL, command=tree.xview)
            tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)
            
            # 布局
            tree.grid(row=0, column=0, sticky='nsew')
            y_scroll.grid(row=0, column=1, sticky='ns')
            x_scroll.grid(row=1, column=0, sticky='ew')
            
            detail_window.grid_rowconfigure(0, weight=1)
            detail_window.grid_columnconfigure(0, weight=1)
            
            # 从数据库加载数据
            try:
                # 获取分析记录ID
                records = self.db.get_analysis_list()
                analysis_id = None
                for rec in records:
                    if rec[1] == record[1] and str(rec[3]) == record[0]:  # 匹配文件名和时间
                        analysis_id = rec[0]
                        break
                
                if analysis_id:
                    requests = self.db.get_requests_by_analysis(analysis_id)
                    for req in requests:
                        tree.insert('', 'end', values=req)
            
            except Exception as e:
                messagebox.showerror("错误", f"加载请求数据失败: {str(e)}")
            
        except Exception as e:
            messagebox.showerror("错误", f"显示记录详情失败: {str(e)}")

    def export_selected_records(self):
        """导出选中的记录到Excel"""
        try:
            # 确保有选中的项目
            if not self.db_tree.selection():
                messagebox.showwarning("警告", "请先选择要导出的记录")
                return
            
            # 选择保存路径
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel文件", "*.xlsx")],
                title="导出分析报告"
            )
            if not filename:
                return
            
            # 导出数据
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "分析结果"
            
            # 写入表头
            headers = ['时间', '文件名', '请求数', 'Hash值']
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # 写入数据
            for row, item_id in enumerate(self.db_tree.selection(), 2):
                values = self.db_tree.item(item_id)['values']
                for col, value in enumerate(values, 1):
                    ws.cell(row=row, column=col, value=value)
            
            # 保存文件
            wb.save(filename)
            messagebox.showinfo("成功", "数据已导出到Excel文件")
            
        except Exception as e:
            messagebox.showerror("错误", f"导出Excel失败: {str(e)}")

    def delete_selected_records(self):
        """删除选中的数据库记录"""
        selected_items = self.db_tree.selection()
        if not selected_items:
            return
        
        if messagebox.askyesno("确认", f"确定要删除选中的 {len(selected_items)} 条记录吗？"):
            success_count = 0
            failed_count = 0
            
            for item_id in selected_items:
                try:
                    # 获取记录信息
                    values = self.db_tree.item(item_id)['values']
                    filename = values[1]
                    analysis_time = values[0]
                    
                    # 从数据库中删除
                    if self.db.delete_record(filename, analysis_time):
                        # 从显示中删除
                        self.db_tree.delete(item_id)
                        success_count += 1
                    else:
                        failed_count += 1
                        
                except Exception as e:
                    print(f"删除记录时出错: {str(e)}")
                    failed_count += 1
                    continue
            
            # 显示删除结果
            if failed_count == 0:
                messagebox.showinfo("成功", f"成功删除 {success_count} 条记录")
            else:
                messagebox.showwarning("完成", 
                    f"删除操作完成\n"
                    f"成功: {success_count} 条\n"
                    f"失败: {failed_count} 条")

    def select_all_records(self):
        """选择所有记录"""
        self.db_tree.selection_set(self.db_tree.get_children())

    def push_frida_server(self):
        """推送并启动 frida-server"""
        try:
            if self.analyzer.push_frida_server():
                messagebox.showinfo("成功", "frida-server 已成功推送并启动")
            else:
                messagebox.showerror("错误", "frida-server 推送或启动失败")
        except Exception as e:
            messagebox.showerror("错误", f"操作失败: {str(e)}")

    def verify_frida_server(self):
        """验证 frida-server 状态"""
        try:
            if self.analyzer.verify_frida_server():
                messagebox.showinfo("成功", "frida-server 运行正常")
            else:
                messagebox.showerror("错误", "frida-server 未正常运行")
        except Exception as e:
            messagebox.showerror("错误", f"验证失败: {str(e)}")

    def show_db_menu(self, event):
        """显示数据库表格的右键菜单"""
        try:
            # 确保有选中的项目
            if self.db_tree.selection():
                self.db_menu.tk_popup(event.x_root, event.y_root)
        finally:
            self.db_menu.grab_release()

if __name__ == "__main__":
    app = ApkAnalyzerGUI()
    app.run() 