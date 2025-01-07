import os
import json
import openpyxl
from datetime import datetime
from urllib.parse import urlparse

class OutputHandler:
    def __init__(self, output_dir):
        self.output_dir = output_dir
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
            
    def save_results(self, results):
        """保存分析结果"""
        try:
            # 获取APK基本信息
            base_name = os.path.splitext(os.path.basename(results['file_name']))[0]
            timestamp = datetime.now()
            
            # 统计网络请求信息
            requests = results.get('requests', [])
            stats = {
                'total': len(requests),
                'domains': len(set(req.get('domain', '') for req in requests if req.get('domain'))),
                'ips': len(set(req.get('ip', '') for req in requests if req.get('ip') and req.get('ip') != 'Unknown')),
                'types': {}
            }
            
            # 统计不同类型的请求数量
            for req in requests:
                req_type = req.get('type', 'Unknown')
                stats['types'][req_type] = stats['types'].get(req_type, 0) + 1
            
            # 构建类型统计字符串
            type_stats = []
            for req_type, count in stats['types'].items():
                type_stats.append(f"{req_type[:3]}{count}")
            
            # 构建文件名
            # 格式: APK名称_日期_时间_[类型统计]_D域名数_I地址数_R总数
            file_prefix = (
                f"{base_name}_"
                f"{timestamp.strftime('%Y%m%d_%H%M%S')}_"
                f"[{'-'.join(type_stats)}]_"
                f"D{stats['domains']}_"
                f"I{stats['ips']}_"
                f"R{stats['total']}"
            )
            
            # 如果文件名太长，使用简短版本
            if len(file_prefix) > 150:
                # 截断APK名称
                if len(base_name) > 30:
                    base_name = base_name[:27] + "..."
                
                # 使用简短格式
                file_prefix = (
                    f"{base_name}_"
                    f"{timestamp.strftime('%Y%m%d_%H%M%S')}_"
                    f"D{stats['domains']}_"
                    f"I{stats['ips']}_"
                    f"R{stats['total']}"
                )
            
            # 保存JSON结果
            json_path = os.path.join(self.output_dir, f"{file_prefix}.json")
            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump(results, f, indent=4, ensure_ascii=False)
            
            # 保存Excel结果
            excel_path = os.path.join(self.output_dir, f"{file_prefix}.xlsx")
            self._save_excel(results, excel_path)
            
            # 打印分析报告
            print("\n=== 分析报告 ===")
            print(f"APK文件: {base_name}")
            print(f"分析时间: {timestamp.strftime('%Y-%m-%d %H:%M:%S')}")
            print("\n网络请求统计:")
            print(f"总请求数: {stats['total']}")
            print(f"域名数量: {stats['domains']}")
            print(f"IP地址数: {stats['ips']}")
            print("\n请求类型分布:")
            for req_type, count in stats['types'].items():
                print(f"- {req_type}: {count}")
            print("\n保存位置:")
            print(f"Excel文件: {os.path.basename(excel_path)}")
            print(f"JSON文件: {os.path.basename(json_path)}")
            print(f"目录: {self.output_dir}")
            print("=" * 20 + "\n")
            
            return {
                'json_path': json_path,
                'excel_path': excel_path,
                'stats': stats,
                'timestamp': timestamp.isoformat()
            }
            
        except Exception as e:
            print(f"保存结果时出错: {str(e)}")
            print(f"APK文件: {results.get('file_name', 'unknown')}")
            print(f"输出目录: {self.output_dir}")
            raise 

    def _format_timestamp(self, timestamp):
        """格式化时间戳"""
        try:
            if not timestamp:  # 处理空值
                return ''
            
            if isinstance(timestamp, str):
                # 处理ISO格式时间戳
                if 'Z' in timestamp:
                    from datetime import datetime
                    dt = datetime.fromisoformat(timestamp.replace('Z', '+00:00'))
                    return dt.strftime('%Y-%m-%d %H:%M:%S')
                # 处理Unix时间戳字符串
                elif timestamp.replace('.', '').isdigit():
                    from datetime import datetime
                    try:
                        # 尝试处理毫秒时间戳
                        ts = float(timestamp)
                        if ts > 1e10:  # 假设是毫秒时间戳
                            ts = ts / 1000
                        return datetime.fromtimestamp(ts).strftime('%Y-%m-%d %H:%M:%S')
                    except:
                        return timestamp
                # 其他字符串格式，直接返回
                return timestamp
            elif isinstance(timestamp, (int, float)):
                # 处理数字类型的时间戳
                from datetime import datetime
                try:
                    if timestamp > 1e10:  # 假设是毫秒时间戳
                        timestamp = timestamp / 1000
                    return datetime.fromtimestamp(timestamp).strftime('%Y-%m-%d %H:%M:%S')
                except:
                    return str(timestamp)
            else:
                # 其他类型，转为字符串
                return str(timestamp)
        except Exception as e:
            print(f"格式化时间戳出错: {str(e)}, 时间戳值: {timestamp}, 类型: {type(timestamp)}")
            return str(timestamp)

    def _save_excel(self, results, excel_path):
        """保存为Excel格式"""
        try:
            print("开始保存Excel...")
            
            wb = openpyxl.Workbook()
            
            # 创建基本信息sheet
            ws_info = wb.active
            ws_info.title = "基本信息"
            
            # 写入APK基本信息
            info_headers = ['项目', '值']
            for col, header in enumerate(info_headers, 1):
                cell = ws_info.cell(row=1, column=col)
                cell.value = header
                cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = openpyxl.styles.PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            
            # 获取统计信息
            requests = results.get('requests', [])
            domains = len(set(req.get('domain', '') for req in requests if req.get('domain')))
            ips = len(set(req.get('ip', '') for req in requests if req.get('ip') and req.get('ip') != 'Unknown'))
            
            info_data = [
                ['APK文件名', results.get('file_name', '')],
                ['文件Hash', results.get('hash', '')],
                ['分析时间', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
                ['总请求数', len(requests)],
                ['域名数量', domains],
                ['IP地址数', ips]
            ]
            
            for row, (item, value) in enumerate(info_data, 2):
                ws_info.cell(row=row, column=1, value=item)
                ws_info.cell(row=row, column=2, value=value)
            
            # 调整基本信息表格列宽
            ws_info.column_dimensions['A'].width = 20
            ws_info.column_dimensions['B'].width = 50
            
            # 创建网络请求sheet
            ws_requests = wb.create_sheet("网络请求详情")
            
            # 定义表头和列宽
            columns = [
                ('时间', 20), 
                ('请求类型', 15),
                ('子类型', 15),
                ('域名', 30),
                ('IP地址', 15),
                ('端口', 10),
                ('URL', 50),
                ('请求方法', 10),
                ('协议', 10),
                ('国家', 15),
                ('地区', 15),
                ('城市', 15),
                ('ISP', 20),
                ('组织', 20)
            ]
            
            # 写入表头
            for col, (header, width) in enumerate(columns, 1):
                cell = ws_requests.cell(row=1, column=col)
                cell.value = header
                cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
                cell.fill = openpyxl.styles.PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = openpyxl.styles.Alignment(horizontal="center")
                ws_requests.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
            
            # 写入请求数据
            for row, request in enumerate(requests, 2):
                try:
                    # 处理时间戳
                    timestamp = self._format_timestamp(request.get('timestamp', ''))
                    
                    # 解析URL信息
                    url = request.get('url', '')
                    protocol = ''
                    if url:
                        try:
                            parsed = urlparse(url)
                            protocol = parsed.scheme
                        except:
                            pass
                    
                    # 准备行数据
                    row_data = [
                        timestamp,
                        request.get('type', ''),
                        request.get('subtype', ''),
                        request.get('domain', ''),
                        request.get('ip', ''),
                        request.get('port', ''),
                        url,
                        request.get('method', ''),
                        protocol or request.get('scheme', ''),
                        request.get('country', ''),
                        request.get('region', ''),
                        request.get('city', ''),
                        request.get('isp', ''),
                        request.get('org', '')
                    ]
                    
                    # 写入数据并设置样式
                    for col, value in enumerate(row_data, 1):
                        cell = ws_requests.cell(row=row, column=col)
                        cell.value = str(value) if value is not None else ''
                        cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)
                        
                        # 为不同类型的请求设置不同的背景色
                        if col == 2:  # 请求类型列
                            if value == 'WebSocket':
                                cell.fill = openpyxl.styles.PatternFill(start_color="E6B8B7", end_color="E6B8B7", fill_type="solid")
                            elif value == 'Network':
                                cell.fill = openpyxl.styles.PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
                            elif value == 'WebView':
                                cell.fill = openpyxl.styles.PatternFill(start_color="C4D79B", end_color="C4D79B", fill_type="solid")
                
                except Exception as e:
                    print(f"处理第 {row-1} 行数据时出错: {str(e)}")
                    print(f"请求数据: {request}")
                    continue
            
            # 添加筛选器
            ws_requests.auto_filter.ref = ws_requests.dimensions
            
            # 冻结首行和首列
            ws_requests.freeze_panes = 'B2'
            
            # 保存文件
            wb.save(excel_path)
            print("Excel文件保存成功")
            
        except Exception as e:
            print(f"保存Excel文件时出错: {str(e)}")
            if 'request' in locals():
                print(f"当前处理的请求数据: {request}")
            raise 